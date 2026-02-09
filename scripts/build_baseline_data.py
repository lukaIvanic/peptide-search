from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATASETS_DIR = ROOT / "Peptide LLM" / "Datasets"
OUT_DIR = ROOT / "app" / "baseline" / "data"
OUT_DIR.mkdir(parents=True, exist_ok=True)

BASELINE_SCHEMA_VERSION = "v1"


def _clean(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    lowered = text.lower()
    if lowered in {"n/a", "na", "null", "none"}:
        return None
    return text


def _parse_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(str(value).strip())
    except Exception:
        return None


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")


def _make_case(
    case_id: str,
    dataset: str,
    sequence: Optional[str],
    labels: List[str],
    n_terminal: Optional[str] = None,
    c_terminal: Optional[str] = None,
    doi: Optional[str] = None,
    pubmed_id: Optional[str] = None,
    paper_url: Optional[str] = None,
    pdf_url: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    return {
        "id": case_id,
        "dataset": dataset,
        "sequence": sequence,
        "n_terminal": n_terminal,
        "c_terminal": c_terminal,
        "labels": labels,
        "doi": doi,
        "pubmed_id": pubmed_id,
        "paper_url": paper_url,
        "pdf_url": pdf_url,
        "metadata": metadata or {},
    }


def build_self_assembly() -> List[Dict[str, Any]]:
    path = DATASETS_DIR / "42256_2024_928_MOESM3_ESM.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    cases: List[Dict[str, Any]] = []
    pos_index = 1
    neg_index = 1

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_negative = "non-assembling" in sheet_name.lower()
        for row in ws.iter_rows(min_row=2, values_only=True):
            sequence = _clean(row[0])
            if not sequence:
                continue
            if is_negative:
                doi = _clean(row[1])
                labels = ["non-self-assembly"]
                case_id = f"self_assembly:neg:{neg_index}"
                neg_index += 1
                metadata = {"source_sheet": sheet_name}
            else:
                methods = _clean(row[1])
                doi = _clean(row[2])
                labels = ["self-assembly"]
                case_id = f"self_assembly:pos:{pos_index}"
                pos_index += 1
                metadata = {
                    "source_sheet": sheet_name,
                    "validation_methods_raw": methods,
                }
            cases.append(
                _make_case(
                    case_id=case_id,
                    dataset="self_assembly",
                    sequence=sequence,
                    labels=labels,
                    doi=doi,
                    metadata=metadata,
                )
            )
    return cases


def build_llps() -> List[Dict[str, Any]]:
    path = DATASETS_DIR / "LLPS peptides.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    cases: List[Dict[str, Any]] = []
    index = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        sequence = _clean(row[1])
        if not sequence:
            continue
        llps_flag = _clean(row[6]) or ""
        labels = ["llps"] if llps_flag.strip().lower().startswith("y") else ["llps-negative"]
        metadata = {
            "peptide_name": _clean(row[0]),
            "concentration": _clean(row[2]),
            "temperature_c": row[3],
            "buffer": _clean(row[4]),
            "additives": _clean(row[5]),
            "llps_flag": llps_flag,
        }
        case_id = f"llps:{index}"
        index += 1
        cases.append(
            _make_case(
                case_id=case_id,
                dataset="llps",
                sequence=sequence,
                labels=labels,
                paper_url=_clean(row[7]),
                metadata=metadata,
            )
        )
    return cases


def _read_csv_rows(path: Path) -> List[Dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        reader = csv.DictReader(handle)
        return [row for row in reader]


def build_catalytic_prot() -> List[Dict[str, Any]]:
    path = DATASETS_DIR / "Peptides cat. func. prot. AA.csv"
    rows = _read_csv_rows(path)
    cases: List[Dict[str, Any]] = []
    index = 1
    for row in rows:
        sequence = _clean(row.get("Peptide sequence (one letter code)"))
        if not sequence:
            continue
        metadata = {
            "smiles": _clean(row.get("SMILES format")),
            "substrate": _clean(row.get("Substrate")),
            "kcat_km": _clean(row.get("Kcat/KM (M-1 s-1)")),
            "secondary_structure": _clean(row.get("Secondary Structure/Self-Assembly")),
            "mechanism": _clean(row.get("Mechanism")),
        }
        case_id = f"catalytic_prot:{index}"
        index += 1
        cases.append(
            _make_case(
                case_id=case_id,
                dataset="catalytic_prot",
                sequence=sequence,
                labels=["catalytic"],
                n_terminal=_clean(row.get("N-terminus")),
                c_terminal=_clean(row.get("C-terminus")),
                doi=_clean(row.get("DOI")),
                metadata=metadata,
            )
        )
    return cases


def build_catalytic_non_prot() -> List[Dict[str, Any]]:
    path = DATASETS_DIR / "Peptides cat. func. non-prot. AA.csv"
    rows = _read_csv_rows(path)
    cases: List[Dict[str, Any]] = []
    index = 1
    for row in rows:
        sequence = _clean(row.get("Peptide sequence (one letter code)"))
        if not sequence:
            continue
        metadata = {
            "substrate": _clean(row.get("Substrate")),
            "kcat_km": _clean(row.get("Kcat/KM (M-1 s-1)")),
            "secondary_structure": _clean(row.get("Secondary Structure/Self-Assembly")),
            "mechanism": _clean(row.get("Mechanism")),
            "non_proteinogenic_residues": _clean(row.get("Non-proteinogenic residues")),
        }
        case_id = f"catalytic_non_prot:{index}"
        index += 1
        cases.append(
            _make_case(
                case_id=case_id,
                dataset="catalytic_non_prot",
                sequence=sequence,
                labels=["catalytic"],
                n_terminal=_clean(row.get("N-terminus")),
                c_terminal=_clean(row.get("C-terminus")),
                doi=_clean(row.get("DOI")),
                metadata=metadata,
            )
        )
    return cases


def build_avp() -> List[Dict[str, Any]]:
    path = DATASETS_DIR / "AVP dataset.csv"
    rows = _read_csv_rows(path)
    cases: List[Dict[str, Any]] = []
    for row in rows:
        avp_id = _clean(row.get("AVP ID"))
        sequence = _clean(row.get("Sequence"))
        if not avp_id or not sequence:
            continue
        case_id = f"avp:{avp_id}"
        metadata = {
            "avp_id": avp_id,
            "virus": _clean(row.get("Virus")),
            "length": _parse_int(row.get("Length")) or _clean(row.get("Length")),
        }
        cases.append(
            _make_case(
                case_id=case_id,
                dataset="avp",
                sequence=sequence,
                labels=["antiviral"],
                pubmed_id=_clean(row.get("PubMed/Patent_ID")),
                metadata=metadata,
            )
        )
    return cases


def main() -> None:
    datasets = [
        {
            "id": "self_assembly",
            "label": "Self-assembly (positive + negative)",
            "description": "Self-assembling vs non-assembling peptide sequences with DOI and validation method.",
            "source_file": "Peptide LLM/Datasets/42256_2024_928_MOESM3_ESM.xlsx",
            "builder": build_self_assembly,
        },
        {
            "id": "llps",
            "label": "LLPS peptides",
            "description": "Phase separation dataset with conditions and article links.",
            "source_file": "Peptide LLM/Datasets/LLPS peptides.xlsx",
            "builder": build_llps,
        },
        {
            "id": "catalytic_prot",
            "label": "Catalytic peptides (proteinogenic)",
            "description": "Catalytic peptides with Kcat/KM and mechanistic notes.",
            "source_file": "Peptide LLM/Datasets/Peptides cat. func. prot. AA.csv",
            "builder": build_catalytic_prot,
        },
        {
            "id": "catalytic_non_prot",
            "label": "Catalytic peptides (non-proteinogenic)",
            "description": "Catalytic peptides including non-proteinogenic residues.",
            "source_file": "Peptide LLM/Datasets/Peptides cat. func. non-prot. AA.csv",
            "builder": build_catalytic_non_prot,
        },
        {
            "id": "avp",
            "label": "Antiviral peptides (AVP)",
            "description": "Antiviral peptides with virus and PubMed/Patent IDs.",
            "source_file": "Peptide LLM/Datasets/AVP dataset.csv",
            "builder": build_avp,
        },
    ]

    index = {
        "schema_version": BASELINE_SCHEMA_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "datasets": [],
        "total_cases": 0,
    }

    total = 0
    for entry in datasets:
        cases = entry["builder"]()
        total += len(cases)
        index["datasets"].append({
            "id": entry["id"],
            "label": entry["label"],
            "description": entry["description"],
            "source_file": entry["source_file"],
            "count": len(cases),
            "file": f"{entry['id']}.json",
        })
        _write_json(OUT_DIR / f"{entry['id']}.json", cases)

    index["total_cases"] = total
    _write_json(OUT_DIR / "index.json", index)
    print(f"Baseline data written to {OUT_DIR} ({total} cases)")


if __name__ == "__main__":
    main()
